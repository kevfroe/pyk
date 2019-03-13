# by krowe
# TODO:

import optparse
import openpyxl
import csv
import sys

#-------------------------------------------------
def fatal(msg):
  print("FATAL ERROR: {}".format(msg))
  sys.exit(1)

#-------------------------------------------------
def row_is_empty(row):
  for cell in row:
    if cell.value != None:
      return False
  return True

#-------------------------------------------------
def get_row_number(row):
  return row[1].row

#-------------------------------------------------
def trace(opts, msg):
  if opts.verbose:
    print(msg)

#-------------------------------------------------
def print_row(opts, row):
  trace(opts, "  row {}: {}".format(get_row_number(row), ",".join([str(cell.value) if cell.value != None else "" for cell in row])))

#-------------------------------------------------
def print_num_rows_removed(cnt):
  print(f"Removed {cnt} empty rows")

#-------------------------------------------------
def txt_remove_empty_rows(opts):
  rows_removed = 0

  wb = openpyxl.Workbook() 
  sheet = wb.worksheets[0]
  sheet.title = "Sheet1"

  with open(opts.file_in, "r", encoding='cp1252') as fp: # encoding required to read special character 0x96 in text file
    reader = csv.reader(fp, delimiter="\t")
  
    for row in reader:
      row_parts = []
      for cell in row:
        if opts.strip_cells:
          row_parts.append(cell.strip())
        else:
          row_parts.append(cell)

      if len(row_parts):
        sheet.append(row_parts)
      else:
        rows_removed += 1
  
  wb.save(filename=opts.file_out)

  print_num_rows_removed(rows_removed)

#-------------------------------------------------
def xlsx_remove_empty_rows(opts):
  wb = openpyxl.load_workbook(opts.file_in)

  for sheetname in wb.sheetnames:
    rows_to_remove = []
    sheet = wb[sheetname]
    
    for row in sheet.iter_rows():
      if row_is_empty(row):
        rows_to_remove.append(get_row_number(row))
        
    rows_to_remove = sorted(rows_to_remove, reverse=True)
    for row_num in rows_to_remove:
      sheet.delete_rows(row_num)

  wb.save(filename=opts.file_out)
  
  print_num_rows_removed(len(rows_to_remove))

#-------------------------------------------------
def main():
  parser = optparse.OptionParser()
  parser.add_option("--file-in",  dest="file_in",  help="input *.xlsx FILE",  metavar="FILE_IN")
  parser.add_option("--file-out", dest="file_out", help="output *.xlsx FILE", metavar="FILE_OUT")
  parser.add_option("-v", "--verbose", action="store_true", dest="verbose")
  parser.add_option("--strip-cells", action="store_true", dest="strip_cells")

  (opts, _) = parser.parse_args()

  if not opts.file_out.endswith(".xlsx"):
    fatal("file_out must be *.xlsx file")

  if opts.file_in.endswith(".xlsx"):
    trace(opts, "Reading *.xlsx file")
    xlsx_remove_empty_rows(opts)
  else:
    trace(opts, "Reading *.txt file")
    txt_remove_empty_rows(opts)
  
#-------------------------------------------------
if __name__ == "__main__":
  main()